""" Generate FortiGate Firewall Policies based on a JSON File
This script generate FortiGate Firewall Policies with the linked address
and service objects.
Connection to FortiGate works over the API connection.
The input file fromat is based on json (use the template file)

(see description in the README file)
This program is free software: you can redistribute it and/or modify it under
the terms of the GNU General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later
version.
This program is distributed in the hope that it will be useful, but WITHOUT
ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
FOR A PARTICULAR PURPOSE. See the GNU General Public License for more details.
You should have received a copy of the GNU General Public License along with
this program. If not, see <http://www.gnu.org/licenses/>.
"""

__author__ = "Andy Scherer"
__contact__ = "ofee42@gmail.com"
__copyright__ = "Copyright 2022, Andy Scherer"
# __credits__ = ["None"]
__date__ = "2022/10/18"
__deprecated__ = False
__email__ = "ofee42@gmail.com"
__license__ = "GPLv3"
__maintainer__ = "developer"
__status__ = "Development"
__version__ = "0.2"

####################
# Importer
####################

import json
import os
import pyfiglet
import requests
import re
import time
import urllib3
import urllib.parse
from openpyxl import load_workbook
from strenum import StrEnum
from helper import colors

####################
# Init's
####################
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

USERNAME = ''
BASE_URL = ''
FGT_TOKEN = ''
VDOM = 'root'
HEADER_KOM = {}
HEADER_NET = {}
HEADER_SVC = {}


class RULE_STATUS(StrEnum):
    READY = "Ready for Implementation"
    DONE = "Implemented"


####################
# Functions
####################
def open_excelfile(filepath):
    try:
        print(colors.fg.yellow, "[...]", colors.reset, "Open Excel File", end="\r")
        excel = load_workbook(filename=filepath, data_only=True)
        sheet_kom = excel["Kommunikationslayout UMB DC LAN"]
        sheet_net = excel["Network Groups"]
        sheet_svc = excel["Service Groups"]

        # Read Header Column
        excel_detect_colnames(sheet_kom, sheet_net, sheet_svc)

        print(colors.fg.green, "[+]", colors.reset, "Excel File successfull openend.")
        return sheet_kom, sheet_net, sheet_svc, excel

    except:
        print(colors.fg.red, "[!]", colors.reset, "Excelfile cannot open!")
        print(colors.bg.red, "[!] Solve error and rerun script. bye!")
        # Error found, exit Script
        raise SystemExit


def excel_detect_colnames(tb_kom, tb_net, tb_svc):
    global HEADER_KOM
    global HEADER_NET
    global HEADER_SVC

    # Get Headernames in Kom Sheet
    current = 0
    for COL in tb_kom.iter_cols(1, tb_kom.max_column):
        HEADER_KOM[COL[0].value] = current
        current += 1

    # Get Headernames in Network Sheet
    current = 0
    for COL in tb_net.iter_cols(1, tb_net.max_column):
        HEADER_NET[COL[0].value] = current
        current += 1

    # Get Headernames in Service Sheet
    current = 0
    for COL in tb_svc.iter_cols(1, tb_svc.max_column):
        HEADER_SVC[COL[0].value] = current
        current += 1


def validate_rows(tab_kom, tab_net, tab_svc):
    for row_cells in tab_kom.iter_rows(min_row=2):
        # Check if rule ready for implementation
        if row_cells[HEADER_KOM['Status']].value == RULE_STATUS.READY:
            print(colors.fg.yellow, "[...]", colors.reset, "Row", row_cells[0].row, "- Validating",
                  end="\r")

            # Validate Src and Dst
            validate_excel_src_dst(row_cells[0].row, tab_net, row_cells[HEADER_KOM['Source']].value, "Source")
            validate_excel_src_dst(row_cells[0].row, tab_net, row_cells[HEADER_KOM['Destination']].value, "Destination")

            # Validate Service
            validate_excel_service(row_cells[0].row, tab_svc, row_cells[HEADER_KOM['Port']].value)

            # Check Interfaces
            fgt_get_interface(row_cells[HEADER_KOM['Src Interface']].value)
            fgt_get_interface(row_cells[HEADER_KOM['Dst Interface']].value)

            # Check all other fields
            excel_validate_fields(row_cells[0].row,
                                  row_cells[HEADER_KOM['Insert After Policy [ID]']].value,
                                  row_cells[HEADER_KOM['Policy Name']].value)

    print(colors.bg.green, "[+]", " - ", "Validation Successful",
          colors.reset)


def validate_excel_src_dst(row_id, tab_net, field_addr, direction):
    field_addr = field_addr
    addrs = field_addr.split('\n')
    # loop through all Address objects
    for addr in addrs:
        if not validate_ip_address(addr):
            # not a valid IP-Address, try with "Network Groups"
            name_addrs = excel_ip_name_resolution(addr, tab_net)
            if len(name_addrs) == 0:
                # Address resolution was not success, exit script
                app_exit_error("Row {0}".format(row_id), "{0} as {1} is not valid".format(addr, direction))

            # loop through all returned IP-Addresses
            for name_addr in name_addrs:
                if not validate_ip_address(name_addr):
                    app_exit_error("Row {0}".format(row_id), "{0} as {1} is not valid".format(name_addr, direction))


def validate_excel_service(row_id, tab_svc, field_svc):
    svcs = field_svc.split('\n')

    # loop through all Address objects
    for svc in svcs:
        if not validate_service(svc):
            # not valid Service, start name resolution with "Services Groups"
            svc_rows = excel_service_name_resolution(svc, tab_svc)
            if len(svc_rows) == 0:
                app_exit_error("Row {0}".format(row_id), "{0} as Service is not valid".format(svc))

            # Loop through all returend Services
            for svc_row in svc_rows:
              if not validate_service(svc_row):
                app_exit_error("Row {0}".format(row_id), "{0} as Service is not valid".format(svc))


def loop_rows(tab_kom, tab_net, tab_svc):
    # Excel must be closed now
    input("Please close Komm-Layout Excel now and press Enter when done.")

    for row_cells in tab_kom.iter_rows(min_row=2):
        # Check if rule ready for implementation
        if row_cells[HEADER_KOM['Status']].value == RULE_STATUS.READY:

            print("", "[.]", "Work on line {}".format(row_cells[0].row))
            # prepare address objects
            src_addrs = excel_prepare_address(row_cells[HEADER_KOM['Source']].value,
                                              row_cells[HEADER_KOM['Source FQDN']].value, tab_net)
            dst_addrs = excel_prepare_address(row_cells[HEADER_KOM['Destination']].value,
                                              row_cells[HEADER_KOM['Destination FQDN']].value, tab_net)

            # prepare service objects
            dst_svcs = excel_prepare_service(row_cells[HEADER_KOM['Port']].value, tab_svc)

            # Add firewall rule
            rule_id = excel_prepare_rule(row_cells, src_addrs, dst_addrs, dst_svcs)
            # If 'rule_id' is False, error on creating Firewall policy - exit this loop
            if rule_id is False:
                print(colors.fg.yellow, "[!]", colors.reset, "Row", row_cells[0].row, "- Stop this rule, continue next")
                print(colors.bg.orange, "[+]", "Row", row_cells[0].row, "- Finish with Errors", colors.reset)
                continue

            # Move Policy if needed
            excel_prepare_policy_move(row_cells, rule_id)

            # Save Rule Infos to Excel
            tab_kom.cell(row=row_cells[0].row, column=HEADER_KOM['Ersteller'] + 1).value = USERNAME
            tab_kom.cell(row=row_cells[0].row, column=HEADER_KOM['Policy ID'] + 1).value = rule_id
            tab_kom.cell(row=row_cells[0].row, column=HEADER_KOM['Status'] + 1).value = RULE_STATUS.DONE

            print(colors.bg.green, "[+]", "Row", row_cells[0].row, "- ", "Rule", rule_id, "- Finish Successful",
                  colors.reset)


def excel_validate_fields(row_id, insert_after, policy_name):
    print(colors.fg.yellow, "[...]", colors.reset, "Row", row_id, "- Check Excel Values", end="\r")

    # Check if insert_after is INT
    try:
        # try to convert to int
        int(insert_after)
    except:
        app_exit_error("Row {0}".format(row_id), "Insert after ID is not an Integer")

    # Check Rulename exist
    try:
        if len(policy_name) < 10:
            app_exit_error("Row {0}".format(row_id), "Policy Name is too short (<10 characters)")
    except:
        app_exit_error("Row {0}".format(row_id), "Check Policy Name")


def excel_ip_name_resolution(addr, tab_net):
    # Start name resolution with "Network Groups"
    excl_values = []
    for net_row in tab_net.iter_rows(HEADER_NET['Group']):
        for cell in net_row:
            if cell.value == addr:
                excl_values = net_row[HEADER_NET['Member']].value
                excl_values = excl_values.split('\n')

    return excl_values


def excel_service_name_resolution(svc_name, tab_svc):
    # Start name resolution with "Service Groups"
    excl_values = []
    for svc_row in tab_svc.iter_rows(HEADER_SVC['Group']):
        for cell in svc_row:
            if str(cell.value).lower() == str(svc_name).lower():
                excl_values = svc_row[HEADER_SVC['Member']].value
                excl_values = excl_values.split('\n')

    return excl_values


def excel_prepare_rule(row_cells, src_addr, dst_addr, dst_service):
    print(colors.fg.yellow, "[...]", colors.reset, "Row", row_cells[0].row, "- Adding Rule to Firewall", end="\r")

    description = row_cells[HEADER_KOM['Kommentar']].value
    # check DNAT
    if row_cells[HEADER_KOM['DNAT']].value == "enable":
        nat = "enable"
    else:
        nat = "disable"

    obj_rule = {
        "name": row_cells[HEADER_KOM['Policy Name']].value,
        "comments": f'{time.strftime("%Y%m%d")}/{USERNAME} - {description}',
        "srcintf": [{"name": row_cells[HEADER_KOM['Src Interface']].value}],
        "dstintf": [{"name": row_cells[HEADER_KOM['Dst Interface']].value}],
        "srcaddr": src_addr,
        "dstaddr": dst_addr,
        "service": dst_service,
        "after_policy": int(row_cells[HEADER_KOM['Insert After Policy [ID]']].value),
        "logtraffic": "all",
        "nat": nat,
        "action": "accept",
        "schedule": "always"
    }

    # Add Rule to FortiGate
    result = fgt_post_policy(obj_rule, row_cells[0].row)
    # if 'result' is False, error on creating Rule.
    if result is False:
        return False

    return result['mkey']


def excel_prepare_address(field_addr, field_fqdn, tab_net):
    # Object with all address objects to return
    obj_addresses = []

    addrs = field_addr.split('\n')
    fqdns = {}
    if type(field_fqdn) is str:
        fqdns = field_fqdn.split('\n')

    # loop through all Address objects
    for addr in addrs:
        if validate_ip_address(addr):
            print(colors.fg.yellow, "[...]", colors.reset, addr, "- Check Address Object", end="\r")

            # change 0.0.0.0/0 to any
            if '0.0.0.0/0' in addr:
                addr = 'all'
                obj_addr = excel_create_address(addr, "")
            else:
                # prepare IPs directly
                addr_ind = addrs.index(addr)
                # add FQDNS if available
                if len(fqdns) == len(addrs):
                    obj_addr = excel_create_address(addr, fqdns[addr_ind])
                else:
                    # if have the list not the same amount of entries, dont add FQDNs
                    obj_addr = excel_create_address(addr, "")

            obj_addresses.append(obj_addr)
        else:
            # not valid IP Address, start name resolution with "Network Groups"
            for net_row in tab_net.iter_rows(HEADER_NET['Group']):
                for cell in net_row:
                    if cell.value == addr:
                        excl_values = net_row[HEADER_NET['Member']].value
                        excl_values = excl_values.split('\n')

            # Loop through all Network Group members
            for value in excl_values:
                # change 0.0.0.0/0 to any
                if '0.0.0.0/0' in value:
                    value = 'all'
                print(colors.fg.yellow, "[...]", colors.reset, value, "- Check Address Object", end="\r")

                obj_adr = excel_create_address(value, "")
                obj_addresses.append(obj_adr)
    return obj_addresses


def excel_prepare_policy_move(row_cells, rule_id):
    print(colors.fg.yellow, "[...]", colors.reset, "Rule", rule_id, "- Moving policy", end="\r")

    after_id = row_cells[HEADER_KOM['Insert After Policy [ID]']].value
    if after_id > 0:
        fgt_move_policy(rule_id, "after", int(after_id))
    else:
        print(colors.fg.green, "[+]", colors.reset, "Rule", rule_id, "- No move needed")


def excel_create_address(field_addr, field_fqdn):
    addr_name = ''
    # Don't check "all" because is default on FortiGates
    if field_addr == "all":
        return {
            "name": "all"
        }

    addr = field_addr.split('/')
    if len(addr) == 1:
        addr_name = f'host_{addr[0]}'
    else:
        if addr[1] == '32':
            addr_name = f'host_{addr[0]}'
        else:
            addr_name = f'net_{field_addr}'

    # Check if a valid IP Address
    if validate_ip_address(field_addr):
        # Check if address on FGT already
        if not fgt_get_address(addr_name):
            # Create address object to FGT
            print(colors.fg.yellow, "[...]", colors.reset, addr_name, "- Create Address Object on Firewall", end="\r")
            fgt_post_address(addr_name, field_fqdn)

    address = {
        "name": addr_name
    }

    return address


def excel_prepare_service(field_svc, tab_svc):
    # Object with all address objects to return
    obj_services = []

    svcs = field_svc.split('\n')

    # loop through all Address objects
    for svc in svcs:
        if validate_service(svc):
            print(colors.fg.yellow, "[...]", colors.reset, svc, "- Check Service Object", end="\r")
            # prepare IPs directly
            obj_svc = excel_create_service(svc)
            obj_services.append(obj_svc)
        else:
            # not valid Service, start name resolution with "Services Groups"
            name_svcs = excel_service_name_resolution(svc, tab_svc)

            # Loop through all Network Group members
            for value in name_svcs:
                print(colors.fg.yellow, "[...]", colors.reset, value, "- Check Address Object", end="\r")
                obj_svc = excel_create_service(value)
                obj_services.append(obj_svc)
    return obj_services


def excel_create_service(field_svc):
    service = field_svc

    # Check if a valid Service Object
    if validate_service(service):
        # Check if already exist
        if not fgt_get_service(service):
            # Create Service on Firewall
            print(colors.fg.yellow, "[...]", colors.reset, field_svc, "- Create Service Object on Firewall", end="\r")
            fgt_post_service(service)
    else:
        app_exit_error(field_svc, "Not a valid Service")
    return {
        "name": service.replace("/", "_")
    }


def excel_save(excel, filepath):
    excel.save(filename=filepath)


def validate_service(svc_string):
    # check if Service already exist
    if fgt_get_service(svc_string):
        return True

    # if not check string
    svc = svc_string.split('/')
    if svc[0] not in ['tcp', 'udp']:
        return False

    if '-' not in svc[1]:
        if int(svc[1]) not in range(1, 65535):
            return False

    return True


def validate_ip_address(ip_string):
    pattern = r"^((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)([/][0-3][0-2]?|[/][1-2][0-9]|[/][0-9])?$"

    # check if valid IP address
    return re.search(pattern, ip_string)


def fgt_get_interface(int_name):
    url = BASE_URL + "cmdb/system/interface/{0}?vdom={1}&access_token={2}".format(int_name, VDOM, FGT_TOKEN)
    res = requests.get(url, verify=False)

    if res.status_code == 200:
        #print(colors.fg.green, "[+]", colors.reset, int_name, "- Interface exists")
        return True
    else:
        # Check may if it's a zone
        if fgt_get_zone(int_name):
            return True

        print(colors.fg.red, "[!]", colors.reset, int_name, "- Interface not EXIST!")
        print(colors.bg.red, "[!] Solve error and rerun script. bye!")
        # Error found, exit Script
        raise SystemExit


def fgt_get_zone(zone_name):
    url = BASE_URL + "cmdb/system/zone/{0}?vdom={1}&access_token={2}".format(zone_name, VDOM, FGT_TOKEN)
    res = requests.get(url, verify=False)

    if res.status_code == 200:
        #print(colors.fg.green, "[+]", colors.reset, zone_name, "- Zone exists")
        return True
    else:
        return False


def fgt_get_service(svc_string):
    url = BASE_URL + "cmdb/firewall.service/custom/{0}?vdom={1}&access_token={2}".format(svc_string.replace("/", "_"),
                                                                                         VDOM, FGT_TOKEN)
    res = requests.get(url, verify=False)

    if res.status_code == 200:
        #print(colors.fg.green, "[+]", colors.reset, svc_string, "- Object already exist")
        return True
    else:
        return False


def fgt_get_address(addr_name):
    # Encode IP Name (e.g. /)
    addr_namede = urllib.parse.quote(addr_name, safe='')
    url = BASE_URL + "cmdb/firewall/address/{0}?vdom={1}&access_token={2}".format(addr_namede, VDOM, FGT_TOKEN)
    res = requests.get(url, verify=False)

    if res.status_code == 200:
        #print(colors.fg.green, "[+]", colors.reset, addr_name, "- Object already exist")
        return True
    else:
        return False


def fgt_post_service(svc_string):
    payload = ''
    svc = svc_string.split('/')

    if "tcp" in svc[0]:
        payload = {
            "name": svc_string.replace("/", "_"),
            "protocol": svc[0],
            "tcp-portrange": svc[1]
        }
    elif "udp" in svc[0]:
        payload = {
            "name": svc_string.replace("/", "_"),
            "protocol": svc[0],
            "udp-portrange": svc[1]
        }

    url = BASE_URL + "cmdb/firewall.service/custom?vdom={0}&access_token={1}".format(VDOM, FGT_TOKEN)
    res = requests.post(url, json=payload, verify=False)

    return api_check_result(res, svc_string)


def fgt_post_address(addr_name, comment):
    # split name in desc and ip
    addr = addr_name.split('_')

    if addr[0] == "host":
        payload = {
            "name": addr_name,
            "subnet": addr[1] + "/32",
            "type": "ipmask",
            "comment": comment
        }
    else:
        payload = {
            "name": addr_name,
            "subnet": addr[1],
            "type": "ipmask",
            "comment": comment
        }

    url = BASE_URL + "cmdb/firewall/address?vdom={0}&access_token={1}".format(VDOM, FGT_TOKEN)
    res = requests.post(url, json=payload, verify=False)

    return api_check_result(res, addr_name)


def fgt_post_policy(obj_policy, row):
    url = BASE_URL + "cmdb/firewall/policy?vdom={0}&access_token={1}".format(VDOM, FGT_TOKEN)
    res = requests.post(url, json=obj_policy, verify=False)
    result = json.loads(res.text.replace("\n", ""))

    if res.status_code == 200:
        print(colors.fg.green, "[+]", colors.reset, "Rule", result['mkey'], "- Firewall Policy created!")
        return result
    else:
        print(colors.fg.red, "[!]", colors.reset, "Row", row, "- ERROR creating Firewall Policy:")
        print(result['cli_error'])
        return False


def fgt_move_policy(rule_id, method, other_id):
    url = BASE_URL + "cmdb/firewall/policy/{}?vdom={}&action=move&{}={}&access_token={}".format(rule_id, VDOM, method,
                                                                                                other_id, FGT_TOKEN)
    res = requests.put(url, verify=False)

    if res.status_code == 200:
        print(colors.fg.green, "[+]", colors.reset, "Rule", rule_id, "- Policy moved successfully")
    else:
        print(colors.fg.red, "[!]", colors.reset, "Rule", rule_id, "- ERROR on policy move:")
        print(res.text)


def api_check_result(res, object_name):
    if res.status_code == 200:
        print(colors.fg.green, "[+]", colors.reset, object_name, "- Object successfully created on Firewall")
        return True
    else:
        print(colors.fg.red, "[!]", colors.reset, object_name, "- ERROR creating Object on Firewall:")
        print(res.text)
        print(colors.bg.red, "[!] Solve error and rerun script. bye!")
        # Error found, exit Script
        raise SystemExit


def app_exit_error(obj_name, error):
    print(colors.fg.red, "[!]", colors.reset, obj_name, "-", error)
    print(colors.bg.red, "[!] Solve error and rerun script. bye!")
    # Error found, exit Script
    raise SystemExit


def main():
    print(pyfiglet.figlet_format("KOM-Layout Import"))
    excel_filepath = input("Kommlayout Excel File Path:")
    username = input("Benutzerkürzel (e.g. scy):")
    baseurl = input("FortiGate IP und Port (e.g. 1.1.1.1:1000")
    token = input("API Token")

    global USERNAME
    USERNAME = username
    global BASE_URL
    BASE_URL = "https://" + baseurl
    global FGT_TOKEN
    FGT_TOKEN = token

    tab_kom, tab_net, tab_svc, excel = open_excelfile(excel_filepath)

    # validate ready rules first
    validate_rows(tab_kom, tab_net, tab_svc)

    # work through all excel rows
    loop_rows(tab_kom, tab_net, tab_svc)

    # Save changed Excel
    excel_save(excel, excel_filepath)


if __name__ == '__main__':
    main()
